from openpyxl import Workbook
from openpyxl.styles import numbers


def export_check(text):
    wb = Workbook()
    ws = wb.active
    # ws.title = 'Лист1'
    data = []
    i = 0
    for s in text.split('\n'):
        i += 1
        product, price, amount = s.split('\t')
        formula = '=B' + str(i) + '*C' + str(i)
        data.append([product, float(price), int(amount), formula])
    data.append(['Итого:', '', '', '=SUM(D1:D' + str(i) + ')'])
    for j in range(1, i + 2):
        ws.cell(column=1, row=j).number_format = numbers.FORMAT_GENERAL
        ws.cell(column=1, row=j).value = data[j - 1][0]
        ws.cell(column=2, row=j).number_format = '#,##0.00_-"₽"'
        ws.cell(column=2, row=j).value = data[j - 1][1]
        ws.cell(column=3, row=j).number_format = numbers.FORMAT_NUMBER
        ws.cell(column=3, row=j).value = data[j - 1][2]
        ws.cell(column=4, row=j).number_format = '#,##0.00_-"₽"'
        ws.cell(column=4, row=j).value = data[j - 1][3]
    wb.save('res.xlsx')


export_check('товар 1\t100\t5\nтовар 2\t200\t6\nтовар 3\t7\t500')

