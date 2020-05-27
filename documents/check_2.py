from openpyxl import Workbook
from openpyxl.styles import numbers


def export_check(text):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Лист1'
    sheet = 1
    for chk in text.split('---'):
        if sheet > 1:
            ws = wb.create_sheet('Лист' + str(sheet))
        sheet += 1
        data = dict()
        for s in chk.split('\n'):
            if len(s) < 2:
                continue
            product, price, amount = s.split('\t')
            if product not in data.keys():
                data[product] = [float(price), int(amount)]
            else:
                data[product][1] += int(amount)
        i = 0
        for p in sorted(data.keys()):
            i += 1
            formula = '=B' + str(i) + '*C' + str(i)
            ws.cell(column=1, row=i).number_format = numbers.FORMAT_GENERAL
            ws.cell(column=1, row=i).value = p
            ws.cell(column=2, row=i).number_format = '#,##0.00_-"₽"'
            ws.cell(column=2, row=i).value = data[p][0]
            ws.cell(column=3, row=i).number_format = numbers.FORMAT_NUMBER
            ws.cell(column=3, row=i).value = data[p][1]
            ws.cell(column=4, row=i).number_format = '#,##0.00_-"₽"'
            ws.cell(column=4, row=i).value = formula
        i += 1
        ws.cell(column=1, row=i).number_format = numbers.FORMAT_GENERAL
        ws.cell(column=1, row=i).value = 'Итого'
        ws.cell(column=4, row=i).number_format = '#,##0.00_-"₽"'
        ws.cell(column=4, row=i).value = '=SUM(D1:D' + str(i - 1) + ')'
    wb.save('res.xlsx')

# export_check('яблоки\t100\t5\nтовар 2\t200\t6\nапельсины\t7\t500\nяблоки\t100\t5\n' +
#     '---\n' +
#     'конфеты\t150\t20\nтовар 5\t200\t6\nколбаса\t8\t600\n'+
#     '---\n' +
#     'помидоры\t50\t10\nтовар 8\t220\t16\nтовар 9\t20\t100\n')
